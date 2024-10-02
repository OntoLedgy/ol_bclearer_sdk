from pathlib import Path

import pandas as pd
from bclearer_interop_services.excel_services.object_model.Sheets import (
    Sheets,
)
from openpyxl import (
    Workbook as OpenpyxlWorkbook,
)
from openpyxl import load_workbook


class Workbooks:
    def __init__(
        self,
        file_path: str = None,
        file_extension: str = "xlsx",
    ):
        self.file_path = file_path

        self.sheets = {}

        if file_path:
            if (
                file_extension
                == ".xlsx"
            ):
                self._load_xlsx(
                    file_path,
                )
            elif (
                file_extension == ".xls"
            ):
                self._load_xls(
                    file_path,
                )
            else:
                raise ValueError(
                    f"Unsupported file extension: {file_extension}",
                )
        else:
            self.wb = OpenpyxlWorkbook()

            self.sheets = {
                sheet.title: Sheets(
                    sheet,
                )
                for sheet in self.wb.worksheets
            }

    def _load_xlsx(self, file_path):
        self.wb = load_workbook(
            file_path,
        )

        for sheet in self.wb.worksheets:
            # Remove any empty rows
            self._remove_empty_rows(
                sheet,
            )

            # Add the sheet to the Sheets object
            self.sheets[sheet.title] = (
                Sheets(sheet)
            )

    def _load_xls(self, file_path):
        xls = pd.ExcelFile(
            file_path,
            engine="xlrd",
        )

        for (
            sheet_name
        ) in xls.sheet_names:
            df = xls.parse(
                sheet_name,
                header=0,
            )
            self.sheets[sheet_name] = (
                self._convert_df_to_sheet(
                    df,
                    sheet_name,
                )
            )

    def _convert_df_to_sheet(
        self,
        dataframe,
        sheet_name,
    ):
        """Converts a pandas DataFrame to the internal Sheets object for .xls files."""
        openpyxl_sheet = OpenpyxlWorkbook().create_sheet(
            sheet_name,
        )

        openpyxl_sheet.append(
            dataframe.columns.tolist(),
        )

        for (
            r_idx,
            row,
        ) in dataframe.iterrows():
            openpyxl_sheet.append(
                row.tolist(),
            )
        return Sheets(openpyxl_sheet)

    def _remove_empty_rows(self, sheet):
        """Removes empty rows where all cells are None from the openpyxl Worksheet object."""
        # Iterate over rows in reverse to avoid index shifting issues
        for row in sheet.iter_rows(
            values_only=True,
        ):
            # If the row is completely empty (all None values), remove the row
            if all(
                cell is None
                for cell in row
            ):
                # Get the index of the row to remove
                row_index = (
                    sheet.max_row
                )
                # Delete the row from the sheet
                sheet.delete_rows(
                    row_index,
                    1,
                )

    def save(
        self,
        file_path: str = None,
    ):
        if file_path is None:
            file_path = self.file_path

        file_type = Path(
            file_path,
        ).suffix.lower()
        if file_type == ".xlsx":
            self._save_xlsx(file_path)
        elif file_type == ".xls":
            self._save_xls(file_path)
        else:
            raise ValueError(
                f"Unsupported file extension: {file_type}",
            )

    def _save_xlsx(self, file_path):
        self.wb.save(file_path)

    def _save_xls(self, file_path):
        with pd.ExcelWriter(
            file_path,
            engine="xlwt",
        ) as writer:
            for (
                sheet_name,
                sheet,
            ) in self.sheets.items():
                dataframe = pd.DataFrame(
                    [
                        [
                            cell.value
                            for cell in row
                        ]
                        for row in sheet.sheet.rows
                    ],
                )
                dataframe.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )

    def sheet(self, sheet_name: str):
        if sheet_name in self.sheets:
            return self.sheets[
                sheet_name
            ]
        raise ValueError(
            f"Sheet {sheet_name} does not exist",
        )

    def create_sheet(self, title: str):
        if title in self.sheets:
            raise ValueError(
                f"Sheet {title} already exists",
            )

        openpyxl_sheet = (
            self.wb.create_sheet(
                title=title,
            )
        )
        sheet = Sheets(openpyxl_sheet)
        self.sheets[title] = sheet
        return sheet

    def remove_sheet(
        self,
        sheet_name: str,
    ):
        if (
            sheet_name
            not in self.sheets
        ):
            raise ValueError(
                f"Sheet {sheet_name} does not exist",
            )

        if (
            sheet_name
            in self.wb.sheetnames
        ):
            del self.wb[sheet_name]
        del self.sheets[sheet_name]
