from bclearer_interop_services.excel_services.object_model.Cells import Cells
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet


class Ranges:
    def __init__(
        self,
        sheet: OpenpyxlWorksheet,
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
    ):
        self.sheet = sheet
        self.min_row = min_row
        self.min_col = min_col
        self.max_row = max_row
        self.max_col = max_col

    def __iter__(self):
        for row in self.sheet.iter_rows(
            min_row=self.min_row,
            max_row=self.max_row,
            min_col=self.min_col,
            max_col=self.max_col,
        ):
            yield [Cells(cell) for cell in row]
